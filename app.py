import streamlit as st
from streamlit_option_menu import option_menu
import pandas as pd
import numpy as np
import plotly.express as px
import os
import warnings
warnings.filterwarnings('ignore')
from PIL import Image
import requests
import openpyxl
import base64


st.set_page_config(page_title="KYGWS", page_icon=":sweat_drops:", layout="wide")


@st.cache_data()
def get_base64_of_bin_file(bin_file):
    with open(bin_file, 'rb') as f:
        data = f.read()
    return base64.b64encode(data).decode()

def set_background(png_file):
    bin_str = get_base64_of_bin_file(png_file)
    page_bg_img = '''
    <style>
    .stApp {
    background-image: url("data:image/png;base64,%s");
    background-size: cover;
    }
    </style>
    ''' % bin_str
    st.markdown(page_bg_img, unsafe_allow_html=True)
    return

set_background('/content/drive/MyDrive/KYGWS/bg.png')

warnings.filterwarnings('ignore')

st.markdown("""
        <style>
               .block-container {
                    padding-top: 1rem;
                    padding-bottom: 0rem;
                    padding-left: 5rem;
                    padding-right: 5rem;
                }
        </style>
        """, unsafe_allow_html=True)


st.image("/content/drive/MyDrive/KYGWS/kygws.png", width=800)
image1 = Image.open('/content/drive/MyDrive/KYGWS/TNAULogo.png')
image2 = Image.open('/content/drive/MyDrive/KYGWS/AECRILogo.png')

# Add images side by side in the sidebar
st.markdown(
    """
    <style>
        [data-testid=stSidebar] [data-testid=stImage]{
            text-align: center;
            display: block;
            margin-left: auto;
            margin-right: auto;
            width: 100%;
        }
    </style>
    """, unsafe_allow_html=True
)

with st.sidebar:
    st.markdown("<h1 style='text-align: center; color: grey;'>KYGWS</h1>", unsafe_allow_html=True)
    col1, col2 = st.columns(2)

    with col1:
        st.image(image1, width=150)

    with col2:
        st.image(image2, width=150)
    st.write("Developed by")
    st.write("Tamil Nadu Agricultural University")
    st.write("Agricultural Engineering College & Research Institute, Coimbatore, TamilNadu (ISO 21001:2018 Certified Institute)")
    selected = option_menu(
        menu_title="Main Menu",
        options=["About", "Status Check", "Glance", "Government Orders", "KYGWS Developers"],
        default_index=0,
    )

if selected == "About":
    st.title(f"{selected}")
    st.markdown('<span style="font-size: 28px;">Know Your Ground Water Status</span>', unsafe_allow_html=True)
    st.markdown('<div style="text-align: justify;"font-size: 24px;">Tamil Nadu is a state in India that balances the development of both agriculture and industry. As of March 2024, it consisted of 38 districts, including 1202 firkas. In light of the fluctuations in the groundwater scenario brought about by recharge and extraction, as well as the requirement for adequate, early intervention for the sustainable management of valuable groundwater resources, the government carried out the Dynamic Groundwater Resources Assessment (DGRA). The State Ground and Surface Water Resources Data Centre, Government of Tamil Nadu, provides the DGRA report. The State Level Technical Coordination Committee suggested Revenue Firka earlier Blocks as the unit of assessment for the DGRA in March 2011 with the governments approval. The Firka status is displayed in DGRA reports and is categorised as Safe, Semi-Critical, Critical, Over Exploited, and Saline/ Poor Quality. The Know Your Ground Water Status (KYGWS) Web App offers information on the temporal scale Firka status and district-specific changes in groundwater levels. There was an increase of Safe Firkas between 2011 and 2023, indicating that groundwater management is performing appropriately. A wide range of stakeholders, including the general public, NGOs, government representatives, academic researchers and farmers, can benefit through KYGWS.</div>', unsafe_allow_html=True)

    cols=st.columns(2)
    with cols[0]:
      st.image("/content/drive/MyDrive/KYGWS/stages_firka.png", use_column_width=True)
    with cols[1]:
      st.image("/content/drive/MyDrive/KYGWS/Firka_map.png", use_column_width=True)

if selected == "Status Check":
    st.title(f"{selected}")

    df = pd.read_csv("/content/drive/MyDrive/KYGWS/firka.csv")

    @st.cache_data
    def convert_df(df):
          return df.to_csv(index=False).encode('utf-8')


    csv = convert_df(df)


    cols=st.columns(2)
    with cols[0]:
      district = st.selectbox("District", df["District"].unique())
      filtered_df = df[df["District"] == district]
      division = st.selectbox("Division", filtered_df["Division"].unique())
      filtered_df1 = df[df["Division"] == division]
    with cols[1]:
      taluk = st.selectbox("Taluk", filtered_df1["Taluk"].unique())
      filtered_df2 = df[df["Taluk"] == taluk]
      firka = st.selectbox("Firka", filtered_df2["Firka"].unique())
      #Search for the firka
      rslt_df = df.loc[df['Firka'] == firka]
      rslt_df = rslt_df.set_index('Firka')
    cols=st.columns(2)
    with cols[0]:
        st.write("Status as on March 2011: ", rslt_df.loc[[firka], ['2011']].values[0][0])
        st.write("Status as on March 2013: ", rslt_df.loc[[firka], ['2013']].values[0][0])
        st.write("Status as on March 2017: ", rslt_df.loc[[firka], ['2017']].values[0][0])
    with cols[1]:
        st.write("Status as on March 2020: ", rslt_df.loc[[firka], ['2020']].values[0][0])
        st.write("Status as on March 2022: ", rslt_df.loc[[firka], ['2022']].values[0][0])
        st.write("Status as on March 2023: ", rslt_df.loc[[firka], ['2023']].values[0][0])
    rslt_df_status = rslt_df.loc[[firka], ['2011', '2013', '2017', '2020', '2022', '2023']]
    rslt_df_status=rslt_df_status.replace('NIL', np.NaN)

    cols=st.columns(6)
    with cols[0]:
      status=rslt_df.loc[[firka], ['2011']].values[0][0]
      df_status=status
      if status=="Safe":
        st.image("/content/drive/MyDrive/KYGWS/Safe.png", use_column_width=True)
        st.markdown(
    """
    <div style='text-align: center;'>
        <span style='color: blue; font-weight: bold;'>2011</span>
    </div>
    """,
    unsafe_allow_html=True
)
      if status=="Semi-Critical":
        st.image("/content/drive/MyDrive/KYGWS/Semicritical.png", use_column_width=True)
        st.markdown(
    """
    <div style='text-align: center;'>
        <span style='color: blue; font-weight: bold;'>2011</span>
    </div>
    """,
    unsafe_allow_html=True
)
      if status=="Critical":
        st.image("/content/drive/MyDrive/KYGWS/Critical.png", use_column_width=True)
        st.markdown(
    """
    <div style='text-align: center;'>
        <span style='color: blue; font-weight: bold;'>2011</span>
    </div>
    """,
    unsafe_allow_html=True
)
      if status=="Over-Exploited":
        st.image("/content/drive/MyDrive/KYGWS/Overexploited.png", use_column_width=True)
        st.markdown(
    """
    <div style='text-align: center;'>
        <span style='color: blue; font-weight: bold;'>2011</span>
    </div>
    """,
    unsafe_allow_html=True
)
      if status=="Saline":
        st.image("/content/drive/MyDrive/KYGWS/Saline.png", use_column_width=True)
        st.markdown(
    """
    <div style='text-align: center;'>
        <span style='color: blue; font-weight: bold;'>2011</span>
    </div>
    """,
    unsafe_allow_html=True
)
    with cols[1]:
      status1=rslt_df.loc[[firka], ['2013']].values[0][0]
      df_status1=status1
      if status1=="Safe":
        st.image("/content/drive/MyDrive/KYGWS/Safe.png", use_column_width=True)
        st.markdown(
    """
    <div style='text-align: center;'>
        <span style='color: blue; font-weight: bold;'>2013</span>
    </div>
    """,
    unsafe_allow_html=True
)
      if status1=="Semi-Critical":
        st.image("/content/drive/MyDrive/KYGWS/Semicritical.png", use_column_width=True)
        st.markdown(
    """
    <div style='text-align: center;'>
        <span style='color: blue; font-weight: bold;'>2013</span>
    </div>
    """,
    unsafe_allow_html=True
)
      if status1=="Critical":
        st.image("/content/drive/MyDrive/KYGWS/Critical.png", use_column_width=True)
        st.markdown(
    """
    <div style='text-align: center;'>
        <span style='color: blue; font-weight: bold;'>2013</span>
    </div>
    """,
    unsafe_allow_html=True
)
      if status1=="Over-Exploited":
        st.image("/content/drive/MyDrive/KYGWS/Overexploited.png", use_column_width=True)
        st.markdown(
    """
    <div style='text-align: center;'>
        <span style='color: blue; font-weight: bold;'>2013</span>
    </div>
    """,
    unsafe_allow_html=True
)
      if status1=="Saline":
        st.image("/content/drive/MyDrive/KYGWS/Saline.png", use_column_width=True)
        st.markdown(
    """
    <div style='text-align: center;'>
        <span style='color: blue; font-weight: bold;'>2013</span>
    </div>
    """,
    unsafe_allow_html=True
)
    with cols[2]:
      status2=rslt_df.loc[[firka], ['2017']].values[0][0]
      df_status2=status2
      if status2=="Safe":
        st.image("/content/drive/MyDrive/KYGWS/Safe.png", use_column_width=True)
        st.markdown(
    """
    <div style='text-align: center;'>
        <span style='color: blue; font-weight: bold;'>2017</span>
    </div>
    """,
    unsafe_allow_html=True
)
      if status2=="Semi-Critical":
        st.image("/content/drive/MyDrive/KYGWS/Semicritical.png", use_column_width=True)
        st.markdown(
    """
    <div style='text-align: center;'>
        <span style='color: blue; font-weight: bold;'>2017</span>
    </div>
    """,
    unsafe_allow_html=True
)
      if status2=="Critical":
        st.image("/content/drive/MyDrive/KYGWS/Critical.png", use_column_width=True)
        st.markdown(
    """
    <div style='text-align: center;'>
        <span style='color: blue; font-weight: bold;'>2017</span>
    </div>
    """,
    unsafe_allow_html=True
)
      if status2=="Over-Exploited":
        st.image("/content/drive/MyDrive/KYGWS/Overexploited.png", use_column_width=True)
        st.markdown(
    """
    <div style='text-align: center;'>
        <span style='color: blue; font-weight: bold;'>2017</span>
    </div>
    """,
    unsafe_allow_html=True
)
      if status2=="Saline":
        st.image("/content/drive/MyDrive/KYGWS/Saline.png", use_column_width=True)
        st.markdown(
    """
    <div style='text-align: center;'>
        <span style='color: blue; font-weight: bold;'>2017</span>
    </div>
    """,
    unsafe_allow_html=True
)
    with cols[3]:
      status3=rslt_df.loc[[firka], ['2020']].values[0][0]
      df_status3=status3
      if status3=="Safe":
        st.image("/content/drive/MyDrive/KYGWS/Safe.png", use_column_width=True)
        st.markdown(
    """
    <div style='text-align: center;'>
        <span style='color: blue; font-weight: bold;'>2020</span>
    </div>
    """,
    unsafe_allow_html=True
)
      if status3=="Semi-Critical":
        st.image("/content/drive/MyDrive/KYGWS/Semicritical.png", use_column_width=True)
        st.markdown(
    """
    <div style='text-align: center;'>
        <span style='color: blue; font-weight: bold;'>2020</span>
    </div>
    """,
    unsafe_allow_html=True
)
      if status3=="Critical":
        st.image("/content/drive/MyDrive/KYGWS/Critical.png", use_column_width=True)
        st.markdown(
    """
    <div style='text-align: center;'>
        <span style='color: blue; font-weight: bold;'>2020</span>
    </div>
    """,
    unsafe_allow_html=True
)
      if status3=="Over-Exploited":
        st.image("/content/drive/MyDrive/KYGWS/Overexploited.png", use_column_width=True)
        st.markdown(
    """
    <div style='text-align: center;'>
        <span style='color: blue; font-weight: bold;'>2020</span>
    </div>
    """,
    unsafe_allow_html=True
)
      if status3=="Saline":
        st.image("/content/drive/MyDrive/KYGWS/Saline.png", use_column_width=True)
        st.markdown(
    """
    <div style='text-align: center;'>
        <span style='color: blue; font-weight: bold;'>2020</span>
    </div>
    """,
    unsafe_allow_html=True
)
    with cols[4]:
      status4=rslt_df.loc[[firka], ['2022']].values[0][0]
      df_status4=status4
      if status4=="Safe":
        st.image("/content/drive/MyDrive/KYGWS/Safe.png", use_column_width=True)
        st.markdown(
    """
    <div style='text-align: center;'>
        <span style='color: blue; font-weight: bold;'>2022</span>
    </div>
    """,
    unsafe_allow_html=True
)
      if status4=="Semi-Critical":
        st.image("/content/drive/MyDrive/KYGWS/Semicritical.png", use_column_width=True)
        st.markdown(
    """
    <div style='text-align: center;'>
        <span style='color: blue; font-weight: bold;'>2022</span>
    </div>
    """,
    unsafe_allow_html=True
)
      if status4=="Critical":
        st.image("/content/drive/MyDrive/KYGWS/Critical.png", use_column_width=True)
        st.markdown(
    """
    <div style='text-align: center;'>
        <span style='color: blue; font-weight: bold;'>2022</span>
    </div>
    """,
    unsafe_allow_html=True
)
      if status4=="Over-Exploited":
        st.image("/content/drive/MyDrive/KYGWS/Overexploited.png", use_column_width=True)
        st.markdown(
    """
    <div style='text-align: center;'>
        <span style='color: blue; font-weight: bold;'>2022</span>
    </div>
    """,
    unsafe_allow_html=True
)
      if status4=="Saline":
        st.image("/content/drive/MyDrive/KYGWS/Saline.png", use_column_width=True)
        st.markdown(
    """
    <div style='text-align: center;'>
        <span style='color: blue; font-weight: bold;'>2022</span>
    </div>
    """,
    unsafe_allow_html=True
)
    with cols[5]:
      status5=rslt_df.loc[[firka], ['2023']].values[0][0]
      df_status5=status5
      if status5=="Safe":
        st.image("/content/drive/MyDrive/KYGWS/Safe.png", use_column_width=True)
        st.markdown(
    """
    <div style='text-align: center;'>
        <span style='color: blue; font-weight: bold;'>2023</span>
    </div>
    """,
    unsafe_allow_html=True
)
      if status5=="Semi-Critical":
        st.image("/content/drive/MyDrive/KYGWS/Semicritical.png", use_column_width=True)
        st.markdown(
    """
    <div style='text-align: center;'>
        <span style='color: blue; font-weight: bold;'>2023</span>
    </div>
    """,
    unsafe_allow_html=True
)
      if status5=="Critical":
        st.image("/content/drive/MyDrive/KYGWS/Critical.png", use_column_width=True)
        st.markdown(
    """
    <div style='text-align: center;'>
        <span style='color: blue; font-weight: bold;'>2023</span>
    </div>
    """,
    unsafe_allow_html=True
)
      if status5=="Over-Exploited":
        st.image("/content/drive/MyDrive/KYGWS/Overexploited.png", use_column_width=True)
        st.markdown(
    """
    <div style='text-align: center;'>
        <span style='color: blue; font-weight: bold;'>2023</span>
    </div>
    """,
    unsafe_allow_html=True
)
      if status5=="Saline":
        st.image("/content/drive/MyDrive/KYGWS/Saline.png", use_column_width=True)
        st.markdown(
    """
    <div style='text-align: center;'>
        <span style='color: blue; font-weight: bold;'>2023</span>
    </div>
    """,
    unsafe_allow_html=True
)


    matching_values = []

    wb = openpyxl.load_workbook("/content/drive/MyDrive/KYGWS/GWL_masterdata.xlsx")
    for sheet_name in wb.sheetnames:
      ws = wb[sheet_name]
      for row in ws.iter_rows(min_row=1, max_col=ws.max_column, values_only=True):
          if row[1] == district:  # Column 2 has an index of 1 (0-based index)
            for col_idx, value in enumerate(row[2:], start=3):
                column_name = ws.cell(row=2, column=col_idx).value  # Get the column name (assuming header is in row 1)
                matching_values.append({
                    "name": str(sheet_name)+str(" ")+str(column_name),
                    "value": value,
                    "Ground_Level":0.0
                    })

    df = pd.DataFrame(matching_values, columns=["name", "value", "Ground_Level"])

    df=df.replace('NIL', np.NaN)
    df=df.replace('Nil', np.NaN)
    df=df.replace('nil', np.NaN)
    df=df.replace('NONE', np.NaN)

    fig = px.bar(df, x="name", y="value", color_discrete_sequence=["blue"], title=str(str(district))).update_layout(
      xaxis_title="Year", yaxis_title="GW Level (m BGL)", showlegend=False, xaxis_tickangle=-45, xaxis_dtick=6, yaxis = dict(autorange="reversed") )
    fig.add_hline(y=00)
    st.plotly_chart(fig)



if selected == "Glance":
    st.title(f"{selected}")
    df = pd.read_csv("/content/drive/MyDrive/KYGWS/firka.csv")

    @st.cache_data
    def convert_df(df):
          return df.to_csv(index=False).encode('utf-8')


    csv = convert_df(df)
    df.replace('Nil', np.NaN)

    df1=df[['2011','2013','2017','2020', '2022', '2023']]
    df1=df1.replace('NIL', np.NaN)
    class_counts = df1.apply(pd.Series.value_counts).fillna("NIL")
    class_counts=class_counts.reindex(index=['Saline', 'Over-Exploited','Critical', 'Semi-Critical','Safe'])
    class_counts1=class_counts.T
    fig = px.line(class_counts1, title='Groundwater Extraction Status Trend', color_discrete_sequence=["yellow", "red", "orange", "blue", "green"]).update_layout(
    xaxis_title="Year", yaxis_title="Number of firkas")
    st.plotly_chart(fig)

    tot_firka = df.groupby('District')['Firka'].count().reset_index()
    tot_firka.columns = ['District', 'Firka_count']
    fig4 = px.bar(tot_firka, x='Firka_count', y='District', title='District wise Number of Firka in Tamil Nadu')
    st.plotly_chart(fig4)


if selected == "Government Orders":
    st.title(f"{selected}")
    st.markdown(""" ## G.O. Related to Firka Catagorization
    ### [G.O.(Ms) No.113 dated 09.06.2016](http://www.groundwatertnpwd.org.in/go113_category_firka.pdf)
    ### [G.O.(Ms) No.257 dated 01.10.2018](http://www.groundwatertnpwd.org.in/gw_go_257_2018.pdf)
    ### [G.O.(Ms) No.161 dated 23.10.2019](http://www.groundwatertnpwd.org.in/G.O.(Ms)%20No.%20161Categorization%20of%20Firka%20as%20on%202017.pdf)
    ### [G.O.(Ms) No.155 dated 28.10.2021](http://www.groundwatertnpwd.org.in/G.O._155_%2028.10.2021.pdf)
    ### [G.O.(Ms) No.15 dated 28.03.2023](http://www.groundwatertnpwd.org.in/GO(Ms)%20No%2015%20WR(R1)%20Dept%20dated%2028.03.2023.pdf)
    ### [G.O.(Ms) No.37 dated 07.03.2024](http://www.groundwatertnpwd.org.in/G.O(Ms.)%20No.37%20WRD%20dated%2007.03.2024.pdf)
    """)


if selected == "KYGWS Developers":
    st.title(f"{selected}")
    col1, col2 = st.columns([5,20])
    with col1:
      st.image("/content/drive/MyDrive/KYGWS/Er.Ponnu.png")

    with col2:
      st.markdown('<span style="font-size: 24px;">Er. P.Ponnuchakkammal</span>', unsafe_allow_html=True)
      st.markdown('<div style="text-align: justify;">The developer completed her under graduate degree in B.Tech Energy and Environmental Engineering and Masters in M.Tech Soil and Water Conservation Engineering. She is currently pursuing her Doctoral degree in Agricultural Engineering specializing Soil and Water Conservation Engineering at Agricultural Engineering College and Research Institute, Tamil Nadu Agricultural University, Coimbatore, India. She possesses research experience of 3 years in watershed management. She has published 4 research papers in various national journals. She has completed high-end workshop entitled Integrated Watershed Management: Development and Emerging Trends at Centre for Water Resources Development and Management, Kerala State Council for Science, Technology and Environment of Government of Kerala.</div>', unsafe_allow_html=True)

    col1, col2 = st.columns([5,20])
    with col1:
      st.image("/content/drive/MyDrive/KYGWS/Dr.RA.png")

    with col2:
      st.markdown('<span style="font-size: 24px;">Dr. A. Raviraj</span>', unsafe_allow_html=True)
      st.markdown('<div style="text-align: justify;">The developer is currently serving as Dean of Agricultural Engineering College and Research Institute, Tamil Nadu Agricultural University, Coimbatore, India. He has Area of specialization in Soil, Water and Conservation engineering. He Has 27 years of experience in teaching, research and extension. He Obtained Ph.D. with Commonwealth Split Site Doctoral Fellowship at University of Newcastle upon Tyne, UK. He has guided as chairman for post-graduate students in the M.Tech. degree and Ph.D. degree scholars. He has served as Chief Scientist in the All India Coordinated Research Project on Irrigation Water Management. He has handled about eleven Externally Funded Schemes in collaboration with WWF, India, TAWDEVA, ICAR, Department of Science and Technology, GOI, IWMI, ADB, and British Geological Survey, UK as PI and Co-PI. He has published many research articles in the International and National journals. He published many books and book chapters. He is recognized various honours and awards such as Best Teacher Award, Common-wealth split-site Doctoral scholarship and IWMI Award for best student research project.</div>', unsafe_allow_html=True)

    col1, col2 = st.columns([5,20])
    with col1:
      st.image("/content/drive/MyDrive/KYGWS/Dr.DS.png")

    with col2:
      st.markdown('<span style="font-size: 24px;">Dr. D. Sureshkumar</span>', unsafe_allow_html=True)
      st.markdown('<div style="text-align: justify;">The developer is currently serving as Director at Centre for Agricultural and Rural Development Studies, Tamil Nadu Agricultural University, Coimbatore, India. An Agricultural Economist, deeply committed to sustainable water resource management and agricultural development through evidence-based and policy-oriented research, policies and dialogues. He has about 29 years of experience in conducting agricultural economics research, teaching for Post Graduate students and outreach activities. He has published 9 books as author/co-author, 21 Technical bulletins, 30 book chapters and 88 research papers in various national and international journals. He has developed a computer-based model for watershed impact evaluation, named as WatDIMP. He has received ICAR Best Teacher Award 1999, D.K. Desai Prize Award 2004, Japanese Award for Outstanding Research on Development 2005, Best Researcher Award 2006,2009 Endeavour Executive Award, Dr. R. T. Joshi Foundation Award (2010), Scientist of Excellence in Niche Area of Research (2019), Eigth Professor Ramesh Chandra Agrawal Award 2022. He has completed 38 research projects with national and international level organizations such as SANDEE, GDN, IFS, IWMI and many more.</div>', unsafe_allow_html=True)

    col1, col2 = st.columns([5,20])
    with col1:
      st.image("/content/drive/MyDrive/KYGWS/Dr.BK.png")

    with col2:
      st.markdown('<span style="font-size: 24px;">Dr. Balaji Kannan</span>', unsafe_allow_html=True)
      st.markdown('<div style="text-align: justify;">The developer is currently serving as Professor & Head, Department of Soil and Water Conservation Engineering, AEC & RI, Tamil Nadu Agricultural University, Coimbatore, India. He has an area of specialization in Remote Sensing and GIS & Soil and Water Conservation Engineering. He completed PG diploma in Geo-Informatics at Indian Institute of Remote Sensing, Dehradun and ITC, Netherlands. He has about 18 years of experience in teaching and research. He Obtained special skills in the application of remote sensing and GIS in natural resource management. He is very passionate in learning and then teaching programming and other digital age tools like python, R, Excel, PowerBI to his students. He has published 5 books, 19 bulletins, 19 book chapters and 40+ research papers in various national and international journals. He has 2 design patents. He guided many M.Tech and Ph.D., students.</div>', unsafe_allow_html=True)

    col1, col2 = st.columns([5,20])
    with col1:
      st.image("/content/drive/MyDrive/KYGWS/Dr.CSS.png")

    with col2:
      st.markdown('<span style="font-size: 24px;">Dr. C. S. Sumathi</span>', unsafe_allow_html=True)
      st.markdown('<div style="text-align: justify;">The developer is currently working as Professor (Computer Science), Department of Physical Science and Information Technology, AEC & RI, Tamil Nadu Agricultural University, Coimbatore, India. She is a teacher and researcher in the field of information technology for the past 16 years. She has an area of interest includes computer programming, data mining, image processing, Artificial Intelligence, e-Learning and web designing. She also associated with many institutional development projects and currently, completed two University Research Projects, one in educational research and another developing institutional books repository. She has been awarded best paper awards in an International Conference held at Kuala Lumpur, Malaysia and 6th National Conference on Agricultural Scientific Tamil organized Agricultural Scientific Tamil Society, New Delhi. She has published many books and research papers in various national and international journals.</div>', unsafe_allow_html=True)

    col1, col2 = st.columns([5,20])
    with col1:
      st.image("/content/drive/MyDrive/KYGWS/Dr.KB.png")

    with col2:
      st.markdown('<span style="font-size: 24px;">Dr. K. Boomiraj</span>', unsafe_allow_html=True)
      st.markdown('<div style="text-align: justify;">The developer is currently working as Associate Professor (Environmental Sciences), Department of Environmental Sciences, Tamil Nadu Agricultural University, Coimbatore, India. He has been awarded PhD. Environmental Sciences by Indian Agricultural Research Institute, New Delhi in 2008. He has an area of specialization in Crop Simulation Model, Climate change impact, adaptation and mitigation studies, Atmospheric trace gas emission studies, Tropospheric Ozone studies. In his piece of work, a part had been contributed as visiting scientist in ICRISAT, Hyderabad. He had handled external funded projects such NICRA, DST SERB and ISRO. He has about 12 years of experience. He has published 5 books, 8 book chapters and 26 research papers in various national and international journals. He guided M.Sc students. He Has been awarded with Best Trainee award (2017), AETDS-Environmentalist Award (2021), NESA Scientist of the Year (2022), Scientist Achiever Award (2024).</div>', unsafe_allow_html=True)
  ###
